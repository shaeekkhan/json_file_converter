import openpyxl
import csv
import json
import sys
import os
import pandas as pd
import tkinter as tk
from tkinter import filedialog


def get_value_from_merged_cell(sheet, cell):
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return sheet.cell(merged_range.min_row, merged_range.min_col).value
    return cell.value


def process_excel(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)

    # Detect if workbook has named tables (common in Google Sheets exports)
    if any(sheet.tables for sheet in wb.worksheets):
        print("Detected table formatting - switching to Pandas")
        df = pd.read_excel(file_path, sheet_name=None)
        return {sheet_name: df[sheet_name].dropna(how='all').to_dict('records') for sheet_name in df.keys()}

    # Fallback to OpenPyXL (default)
    data = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        headers = [get_value_from_merged_cell(ws, cell) for cell in ws[1]]
        last_values = {header: None for header in headers}
        sheet_data = []

        for row in ws.iter_rows(min_row=2):
            row_data = {}
            for header, cell in zip(headers, row):
                value = get_value_from_merged_cell(ws, cell)
                if value is None:
                    value = last_values[header]
                else:
                    last_values[header] = value
                if value is not None:
                    row_data[header] = value
            if row_data:
                sheet_data.append(row_data)

        data[sheet] = sheet_data

    return data


def process_csv(file_path):
    with open(file_path, 'r', newline='', encoding='utf-8') as csvfile:
        reader = csv.DictReader(csvfile)
        data = []
        for row in reader:
            row_data = {k: v for k, v in row.items() if v}
            if row_data:
                data.append(row_data)
    return data


def process_text(file_path):
    data = []
    with open(file_path, 'r', encoding='utf-8') as file:
        for line in file:
            line = line.strip()
            if line:
                data.append({"line": line})
    return data


def file_to_json(input_file, output_file):
    _, file_extension = os.path.splitext(input_file)
    file_extension = file_extension.lower()

    if file_extension in ['.xlsx', '.xls']:
        data = process_excel(input_file)  # Now processes all sheets
    elif file_extension == '.csv':
        data = process_csv(input_file)
    elif file_extension in ['.txt', '.log', '.md']:
        data = process_text(input_file)
    else:
        try:
            df = pd.read_excel(input_file, sheet_name=None)  # Pandas handles multiple sheets
            data = {sheet_name: df[sheet_name].to_dict('records') for sheet_name in df.keys()}
        except Exception as e:
            print(f"Error: Unable to process file. {str(e)}")
            return

    with open(output_file, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=4, ensure_ascii=False)

    print(f"Conversion complete. JSON file saved as {output_file}")


def select_file():
    root = tk.Tk()
    root.withdraw()  # Hide the main window
    file_path = filedialog.askopenfilename(title="Select a file",
                                           filetypes=[("Excel files", "*.xlsx;*.xls"),
                                                      ("CSV files", "*.csv"),
                                                      ("Text files", "*.txt;*.log;*.md"),
                                                      ("All files", "*.*")])
    return file_path


if __name__ == "__main__":
    input_file = select_file()
    if not input_file:
        print("No file selected. Exiting...")
        sys.exit(1)

    output_file = input_file + ".json"  # Automatically generate JSON filename

    file_to_json(input_file, output_file)
