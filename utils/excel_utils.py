import openpyxl
import pandas as pd


def get_value_from_merged_cell(sheet, cell):
    for merged_range in sheet.merged_cells.ranges:
        if cell.coordinate in merged_range:
            return sheet.cell(merged_range.min_row, merged_range.min_col).value
    return cell.value


def process_excel(file_path):
    wb = openpyxl.load_workbook(file_path, data_only=True)

    if any(sheet.tables for sheet in wb.worksheets):
        print("Detected table formatting - switching to Pandas")
        df = pd.read_excel(file_path, sheet_name=None)
        return {sheet_name: df[sheet_name].dropna(how='all').to_dict('records') for sheet_name in df.keys()}

    data = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        headers = [get_value_from_merged_cell(ws, cell) for cell in ws[1]]
        last_values = {header: None for header in headers}
        sheet_data = []

        for row in ws.iter_rows(min_row=2):
            row_data = {}
            for header, cell in zip(headers, row):
                value = get_value_from_merged_cell(ws, cell)
                if value is None and last_values[header] is not None:
                    value = last_values[header]
                elif value is not None:
                    last_values[header] = value
                else:
                    continue
                row_data[header] = value
            if row_data:
                sheet_data.append(row_data)

        data[sheet] = sheet_data

    return data
